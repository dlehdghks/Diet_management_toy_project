import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def update_manual():
    wb = openpyxl.load_workbook("사용설명서.xlsx")
    
    # 1. 프로젝트 개요 보강
    ws1 = wb["프로젝트 개요"]
    # 기존 데이터가 중복되지 않게 관리하거나, 필요시 특정 행을 업데이트하는 로직이 좋지만
    # 여기서는 간단히 새로운 정보를 추가합니다.
    ws1.append(["보안 설정", "JWT (JSON Web Token), 비밀번호 해싱(bcrypt)", "추가됨"])
    ws1.append(["인증 흐름", "OAuth2PasswordBearer", "완료"])

    # 2. 백엔드 API 업데이트
    ws2 = wb["백엔드 API"]
    ws2.delete_rows(2, ws2.max_row)
    new_apis = [
        ["/", "GET", "서버 상태 확인 및 환영 메시지", "완료"],
        ["/health", "GET", "서버 헬스 체크", "완료"],
        ["/register", "POST", "신규 회원가입 (비밀번호 해싱 적용)", "완료"],
        ["/login", "POST", "JWT 토큰 발급 로그인", "완료"],
        ["/users/me", "GET", "현재 로그인한 사용자 정보 조회", "완료"],
        ["/users/", "GET", "전체 사용자 목록 조회", "완료"]
    ]
    for api in new_apis:
        ws2.append(api)
    
    # 3. 프론트엔드 구조 시트 업데이트
    ws3 = wb["프론트엔드 구조"]
    ws3.delete_rows(2, ws3.max_row)
    new_frontend_struct = [
        ["src/App.vue", "메인 컴포넌트", "로그인/회원가입 폼 및 대시보드 UI (JWT 처리)"],
        ["src/api/index.ts", "API 모듈", "Axios 인터셉터(토큰 자동 주입) 및 인증 API"],
        ["src/main.ts", "진입점", "Vue 앱 마운트"]
    ]
    for item in new_frontend_struct:
        ws3.append(item)

    # 4. 시스템 아키텍처 시트 업데이트
    if "시스템 아키텍처" not in wb.sheetnames:
        ws4 = wb.create_sheet("시스템 아키텍처")
        ws4.append(["구분", "설명"])
        for cell in ws4[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    else:
        ws4 = wb["시스템 아키텍처"]
        ws4.delete_rows(2, ws4.max_row)
        
    ws4.append(["Backend Stack", "FastAPI, SQLAlchemy, SQLite(dev), JWT, Passlib"])
    ws4.append(["Frontend Stack", "Vue 3, Vite, TypeScript, Axios"])
    ws4.append(["Auth Flow", "JWT Token stored in LocalStorage"])
    ws4.append(["DB Type", "SQLite (sql_app.db)"])

    wb.save("사용설명서.xlsx")
    print("사용설명서.xlsx가 최신 인증 시스템 기반으로 업데이트되었습니다.")

if __name__ == "__main__":
    update_manual()
