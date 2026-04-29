import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def create_manual():
    wb = openpyxl.Workbook()
    
    # 1. 프로젝트 개요 시트
    ws1 = wb.active
    ws1.title = "프로젝트 개요"
    ws1.append(["항목", "내용", "비고"])
    ws1.append(["프로젝트명", "GEMINI 통합 관리 시스템", ""])
    ws1.append(["백엔드 기술 스택", "FastAPI, SQLAlchemy, PostgreSQL, Pydantic", ""])
    ws1.append(["프론트엔드 기술 스택", "Vue 3, Vite, TypeScript, Tailwind CSS(예정)", ""])
    ws1.append(["데이터베이스", "PostgreSQL (psycopg2-binary)", ""])
    
    # 스타일 적용
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 2. 백엔드 API 시트
    ws2 = wb.create_sheet("백엔드 API")
    ws2.append(["엔드포인트", "메서드", "설명", "상태"])
    ws2.append(["/", "GET", "서버 상태 확인 및 환영 메시지", "대기"])
    
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 3. 프론트엔드 구조 시트
    ws3 = wb.create_sheet("프론트엔드 구조")
    ws3.append(["경로", "구성 요소", "설명"])
    ws3.append(["src/App.vue", "메인 컴포넌트", "애플리케이션의 진입점"])
    ws3.append(["src/components/HelloWorld.vue", "기본 컴포넌트", "초기 화면 구성용"])
    
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    wb.save("사용설명서.xlsx")
    print("사용설명서.xlsx 파일이 생성되었습니다.")

if __name__ == "__main__":
    create_manual()
